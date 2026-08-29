from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


with TemporaryDirectory() as temp_dir:
    path = Path(temp_dir) / "catalog.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Catalog"
    worksheet.append(["item", "status"])
    worksheet.append(["Keyboard", "active"])
    worksheet.append(["Mouse", "inactive"])

    table = Table(displayName="CatalogTable", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    validation = DataValidation(
        type="list",
        formula1='"active,inactive"',
        allow_blank=False,
    )
    worksheet.add_data_validation(validation)
    validation.add("B2:B20")
    workbook.save(path)

    reloaded = load_workbook(path)
    sheet = reloaded["Catalog"]
    print(f"tables: {list(sheet.tables.keys())}")
    print(f"validations: {len(sheet.data_validations.dataValidation)}")
    reloaded.close()
