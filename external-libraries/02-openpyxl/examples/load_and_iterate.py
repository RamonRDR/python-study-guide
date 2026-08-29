from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook


with TemporaryDirectory() as temp_dir:
    path = Path(temp_dir) / "orders.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"
    worksheet.append(["order_id", "amount"])
    worksheet.append([101, 25.5])
    worksheet.append([102, 40.0])
    worksheet.append([103, 34.5])
    workbook.save(path)

    reloaded = load_workbook(path, read_only=True, data_only=True)
    sheet = reloaded["Orders"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    print(f"orders: {len(rows)}")
    print(f"total: {sum(amount for _, amount in rows):.2f}")
    reloaded.close()
