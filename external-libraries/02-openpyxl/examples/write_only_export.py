from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook


with TemporaryDirectory() as temp_dir:
    path = Path(temp_dir) / "streamed.xlsx"

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Data")
    worksheet.append(["id", "value"])
    worksheet.append([1, 10])
    worksheet.append([2, 20])
    worksheet.append([3, 30])
    workbook.save(path)

    reloaded = load_workbook(path, read_only=True, data_only=True)
    sheet = reloaded["Data"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    print(f"rows: {len(rows)}")
    print(f"sum: {sum(value for _, value in rows)}")
    reloaded.close()
