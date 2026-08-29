from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook


with TemporaryDirectory() as temp_dir:
    path = Path(temp_dir) / "report.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet.append(["product", "units", "unit_price", "total"])
    worksheet.append(["Notebook", 2, 15.0, "=B2*C2"])
    worksheet.append(["Marker", 5, 4.0, "=B3*C3"])
    workbook.save(path)

    reloaded = load_workbook(path, data_only=False)
    sheet = reloaded["Summary"]
    print(f"sheet: {sheet.title}")
    print(f"rows: {sheet.max_row}")
    print(f"formula: {sheet['D2'].value}")
    reloaded.close()
