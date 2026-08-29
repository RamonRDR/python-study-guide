from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


with TemporaryDirectory() as temp_dir:
    path = Path(temp_dir) / "styled.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    worksheet.append(["item", "amount"])
    worksheet.append(["Service", 1250.5])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet["B2"].number_format = '#,##0.00'
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 14
    workbook.save(path)

    reloaded = load_workbook(path)
    sheet = reloaded["Report"]
    print(f"header bold: {sheet['A1'].font.bold}")
    print(f"number format: {sheet['B2'].number_format}")
    print(f"freeze panes: {sheet.freeze_panes}")
    reloaded.close()
